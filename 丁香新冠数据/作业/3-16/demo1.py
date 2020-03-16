#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/3/16 12:47
# @Author  : lihanhan
# @Email   : demo1li@163.com
# @File    : demo1.py
import requests
import json
import re
from pyecharts.charts import Bar
from openpyxl import Workbook
from pyecharts.charts import Line, Pie, Map, Page
from pyecharts import options as opts
from pyecharts.globals import ThemeType


def get_html(month, day):  # 获取html资源 并解析为json数据 写入json文件
    url = 'http://3g.dxy.cn/newh5/view/pneumonia'
    response = requests.get(url)
    html = str(response.content, 'UTF-8')
    html_file = open('htmls/%d%d.html' % (month, day), 'w', encoding='UTF-8')
    html_file.write(html)
    html_file.close()
    # 省份数据
    json_file = open('jsons/%d%d.json' % (month, day), 'w', encoding='UTF-8')
    matches = re.findall('\[[^>]+\]', html)
    for match in matches:
        try:
            json_array = json.loads(match)
            json_object = json_array[0]
            if 'provinceName' in json_object and json_object['provinceName'] == '湖北省':
                json_file.write(match)
                break
        except:
            continue
    json_file.close()
    # 总体统计数据 写入json文件
    total_statistic_file = open('jsons/%d%d-总计.json' % (month, day), 'w', encoding='UTF-8')
    matches = re.findall('\{"id":1,[^(>})]+\}', html)
    for match in matches:
        if "infectSource" in match:
            index = match.find(',"marquee":[')
            if index != -1:
                match = match[:index] + '}'
            total_statistic_file.write(match)
            break
    total_statistic_file.close()


def get_total_statistic(month, day):
    file = open('jsons/%d%d-总计.json' % (month, day), 'r', encoding='UTF-8')
    json_object = json.loads(file.read())
    file.close()
    # print(json_object)
    return json_object


def create_excel(month, day):  # 需求1 获得每天的确诊，治愈和死亡人数 保存到Excel文件中 格式为xlsx
    info = []
    total_json = get_total_statistic(month, day)
    currentConfirmedIncr = total_json["currentConfirmedIncr"]  # 确诊人数
    curedIncr = total_json["curedIncr"]  # 治愈人数
    deadIncr = total_json["deadIncr"]  # 死亡人数
    info.append(currentConfirmedIncr)
    info.append(curedIncr)
    info.append(deadIncr)
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = '确诊人数'
    ws['B1'].value = '治愈人数'
    ws['C1'].value = '死亡人数'
    ws.append(info)
    wb.save("today_info.xlsx")


def get_statistic(month, day):
    file = open('jsons/%d%d.json' % (month, day), 'r', encoding='UTF-8')
    json_object = json.loads(file.read())
    file.close()
    return json_object


def count_total_num(month, day):  # 实现数据可视化之柱状图
    List = []  # x轴的名字
    List1 = []  # 湖北的数据
    List2 = []  # 湖北以外的数据
    info_list = []  # 各省份的集合
    json_object = get_statistic(month, day)
    hubei_num_info = json_object[0]  # 湖北数据 现在确诊的人数
    # 获取各省份数据  存入list集合
    for p in json_object:
        provinces = p['provinceShortName']
        confirmedCount = p['confirmedCount']
        info_list.append([provinces, confirmedCount])
    # print(hubei_num_info)
    currentConfirmedCount = hubei_num_info["currentConfirmedCount"]  # 湖北现存确诊的人数
    confirmedCount = hubei_num_info["confirmedCount"]  # 湖北累计确诊的人数
    curedCount = hubei_num_info["curedCount"]  # 湖北累计自愈的人数
    deadCount = hubei_num_info["deadCount"]  # 湖北死亡的人数

    total_json = get_total_statistic(month, day)  # 全国数据
    currentConfirmedCount1 = total_json["currentConfirmedCount"]  # 全国现存确诊
    confirmedCount1 = total_json["confirmedCount"]  # 全国累计确诊
    curedCount1 = total_json["curedCount"]  # 全国累计自愈的人数
    deadCount1 = total_json["deadCount"]  # 全国死亡的人数
    """
    计算
    """
    Count1 = currentConfirmedCount1 - currentConfirmedCount  # 湖北省以外的现存确诊的人数
    Count2 = confirmedCount1 - confirmedCount  # 湖北省以外的累计确诊
    Count3 = curedCount1 - curedCount  # 湖北省以外的累计自愈的人数
    Count4 = deadCount1 - deadCount  # 湖北省以外的死亡的人数
    List2.append(Count1)
    List2.append(Count2)
    List2.append(Count3)
    List2.append(Count4)

    List1.append("现存确诊的人数")
    List1.append("累计确诊的人数")
    List1.append("累计自愈的人数")
    List1.append("死亡的人数")

    List.append(currentConfirmedCount)
    List.append(confirmedCount)
    List.append(curedCount)
    List.append(deadCount)
    bar = (
        Bar(init_opts=opts.InitOpts(theme=ThemeType.LIGHT))
            .add_xaxis(List1)
            .add_yaxis("湖北省", List)
            .add_yaxis("湖北省外", List2)
            .set_global_opts(title_opts=opts.TitleOpts(title="主标题", subtitle="副标题"))
    )
    bar.render('新冠数据柱状图展示.html')


def get_province_data(month, day):
    file = open('jsons/%d%d.json' % (month, day), 'r', encoding='UTF-8')
    json_array = json.loads(file.read())
    file.close()
    sort_arr = [json_obj for json_obj in json_array]
    return sorted(sort_arr, key=lambda x: x['confirmedCount'], reverse=True)


def get_province_status(month, day):
    json_array = get_province_data(month, day)
    data = []
    for province in json_array:
        data.append((province['provinceShortName'], province['confirmedCount']))
    data.sort(key=lambda x: x[1], reverse=True)
    labels = [d[0] for d in data]
    counts = [d[1] for d in data]
    return labels, counts


def get_default_pieces():  # 设置地图不同数量  不同颜色
    return [
        {'min': 1000, 'color': '#450704'},
        {'max': 999, 'min': 100, 'color': '#75140B'},
        {'max': 99, 'min': 10, 'color': '#AD2217'},
        {'max': 9, 'min': 1, 'color': '#DE605B'},
        {'max': 0, 'color': '#FFFEE7'},
    ]


"""
map 函数 地图主要用于地理区域数据的可视化。
map.add(name, attr, value, maptype='china', is_roam=True, is_map_symbol_show=True, **kwargs)
"""


def get_map(labels, counts, where, title, size, pieces):
    my_map = Map(init_opts=opts.InitOpts(width=size[0], height=size[1]))
    my_map.add("新冠数据", [list(z) for z in zip(labels, counts)], where)
    my_map.set_series_opts(label_opts=opts.LabelOpts(font_size=8))
    my_map.set_global_opts(
        title_opts=opts.TitleOpts(title=title),
        legend_opts=opts.LegendOpts(is_show=False),
        visualmap_opts=opts.VisualMapOpts(
            pieces=pieces,
            is_piecewise=True,
            is_show=False
        ),
    )
    return my_map


def draw_multiple_map(month, day):
    page = Page(layout=Page.SimplePageLayout)
    size = ('1000px', '1000px')

    # 全国疫情地图
    labels, counts = get_province_status(month, day)
    country_map = get_map(labels=labels, counts=counts,
                          title='全国-%s例' % get_total_statistic(month, day)['confirmedCount'], size=size, where='china',
                          pieces=get_default_pieces())
    page.add(country_map)
    page.render('全国新冠数据地理分布展示.html')


if __name__ == '__main__':
    m, d = 3, 16
    get_html(m, d)
    create_excel(m, d)
    # count_total_num(m,d)
    draw_multiple_map(m, d)
