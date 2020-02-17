import csv
import time
import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import xlwt


url = 'https://free-api.heweather.net/s6/weather/forecast?location=信阳&key=3f8659a2ab674ca6bdc3e740462a6ae1'
# 获取当日时间	2020-2-17
today_time = time.strftime('%Y-%m-%d', time.localtime(time.time()))
def get_weather_data():

    res = requests.get(url).json()
    result = res['HeWeather6'][0]['daily_forecast']

    city = res['HeWeather6'][0]['basic']['parent_city']#+res['HeWeather6'][0]['location']

    names = ['城市','时间','天气状况','最高温','最低温','日出','日落']
    """
    csv格式
    #      with open('today_weather.csv', 'w', newline='')as f:
    #     writer = csv.writer(f)
    #     writer.writerow(names)
    。。。。。。。
    
    #writer.writerows([(city, date, cond, max, min, sr, ss)])#writerows 加s的是写内容 不加 的是写标题
    """

    """
    xlsx格式
    wb = Workbook()
    ws = wb.active
    # 设置单元格的值，A1等于6(测试可知openpyxl的行和列编号从1开始计算)，B1等于7
    ws.cell(row=1, column=1).value = '城市'
    ws['B1'].value = '时间'
    ws['C1'].value = '天气状况'
    ws['D1'].value = '最高温度'
    ws['E1'].value = '最低温度'
    ws['F1'].value = '日出时间'
    ws['G1'].value = '日落时间'
    ......
            ws.append([city,date,cond,max,min,sr,ss])
        wb.save("today_weather.xlsx")
    
    """

    for data in result:
        date = data['date']
        cond = data['cond_txt_d']
        max = data['tmp_max']
        min = data['tmp_min']
        sr = data['sr']
        ss = data['ss']
        wb = Workbook()
        ws = wb.active
        # 设置单元格的值，A1等于6(测试可知openpyxl的行和列编号从1开始计算)，B1等于7
        ws.cell(row=1, column=1).value = '城市'
        ws['B1'].value = '时间'
        ws['C1'].value = '天气状况'
        ws['D1'].value = '最高温度'
        ws['E1'].value = '最低温度'
        ws['F1'].value = '日出时间'
        ws['G1'].value = '日落时间'
        ws.append([city,date,cond,max,min,sr,ss])
        wb.save("today_weather.xls")
        # info = [
        #     [city,date,cond,max,min,sr,ss]
        # ]
        # book = xlwt.Workbook()
        # sheet = book.add_sheet('xiangxin')
        # # 写入表头
        # i = 0
        # for j in names:
        #     sheet.write(0, i, j)
        #     i += 1
        # # 写入表内容
        # l = 1
        # for d in info:
        #     c = 0
        #     for dd in d:
        #         sheet.write(l, c, dd)
        #         c += 1
        #     l += 1
        # # 保存
        # book.save('today_weather.xls')


    print(city,date,cond,max,min,sr,ss)
    send_email()
def send_email():
    # 设置邮箱的域名
    HOST = 'smtp.qq.com'
    # 设置邮件标题
    SUBJECT = '主子，今日份天气预报到了哟'
    # 设置发件人邮箱
    FROM = '320783214@qq.com'
    # 设置收件人邮箱
    TO = '2530112392@qq.com'	# 可以填写多个邮箱，用逗号分隔，后面会用split做逗号分割
    message = MIMEMultipart('related')
    # --------------------------------------发送文本-----------------
    # 发送邮件正文到对方的邮箱中
    message_html = MIMEText("%s日份天气预报到账啦，请查收(打开下面的附件)" % today_time, 'plain', 'utf-8')	# \n为换行
    message.attach(message_html)

    # -------------------------------------添加文件---------------------
    # 要确定当前目录有test.csv这个文件
    message_xlsx = MIMEText(open('today_weather.xls', 'rb').read(), 'base64', 'utf-8')
    # 设置文件在附件当中的名字
    message_xlsx['Content-Disposition'] = 'attachment;filename="today_weather.xls"'
    message.attach(message_xlsx)

    # 设置邮件发件人
    message['From'] = FROM
    # 设置邮件收件人
    message['To'] = TO
    # 设置邮件标题
    message['Subject'] = SUBJECT

    # 获取简单邮件传输协议的证书
    email_client = smtplib.SMTP_SSL(host='smtp.qq.com')
    # 设置发件人邮箱的域名和端口，端口为465
    email_client.connect(HOST, '465')
    # ---------------------------邮箱授权码------------------------------
    result = email_client.login(FROM, 'vazakcfvsmgnbhif')
    print('登录结果', result)

    email_client.sendmail(from_addr=FROM, to_addrs=TO.split(','), msg=message.as_string())
    # 关闭邮件发送客户端
    email_client.close()
if __name__ == '__main__':
    get_weather_data()

